# -*- coding: UTF-8 -*-
"""
a_io.py
=======

Script :   a_io.py

Author :   Dan.Patterson@carleton.ca

Modified : 2018-12-30

Purpose : Basic io tools for numpy arrays and arcpy

Notes :
::
    1.  load_npy    - load numpy npy files
    2.  save_npy    - save array to *.npy format
    3.  load_txt    - read array created by save_txtt
    4.  save_txt    - save array to npy format
    5.  arr_json    - save to json format
    6-9 dict<->array conversions
    10. excel_np    - convert xls/xlsx files to numpy structured/recarray

---------------------------------------------------------------------
"""
# pylint: disable=C0103
# pylint: disable=R1710
# pylint: disable=R0914

# ---- imports, formats, constants ----
import sys
import numpy as np


ft = {'bool': lambda x: repr(x.astype(np.int32)),
      'float_kind': '{: 0.3f}'.format}
np.set_printoptions(edgeitems=10, linewidth=80, precision=2, suppress=True,
                    threshold=100, formatter=ft)
np.ma.masked_print_option.set_display('-')  # change to a single -

script = sys.argv[0]  # print this should you need to locate the script

__all__ = ['dtype_info',
           'load_npy',
           'save_npy',
           'load_txt',
           'save_txt',
           'arr_json',
           'dict_arrays',
           'iterable_dict',
           'dict_struct',
           'struct_dict',
           'excel_np'
           ]


def dtype_info(a=None, as_string=False):
    """Return dtype information for a structured/recarray
    """
    dt = a.dtype.descr
    names = [i[0] for i in dt]
    formats = [i[1] for i in a.dtype.descr]
    if as_string:
        names = ", ".join([i[1] for i in names])
        formats = ", ".join([i[1] for i in formats])
    return names, formats


# ----------------------------------------------------------------------
# (1) load_npy .... code section ---
def load_npy(f_name, all_info=False):
    """load a well formed `npy` file representing a structured array

    Returns
    -------
        The array, the description, field names and their size.
    """
    a = np.load(f_name)
    if all_info:
        desc = a.dtype.descr
        nms = a.dtype.names
        sze = [i[1] for i in a.dtype.descr]
        return a, desc, nms, sze
    #
    return a


# ----------------------------------------------------------------------
# (2) read_npy .... code section ---
def save_npy(a, f_name):
    """Save an array as an npy file.

    The type of data in each column is arbitrary.  It will be cast to the
    given dtype at runtime
    """
    np.save(f_name, a)


# ----------------------------------------------------------------------
# (3) read_txt .... code section ---
def load_txt(name="arr.txt"):
    """Read the structured/recarray created by save_txt.

    dtype : data type
        If `None`, it allows the structure to be read from the array.

    delimiter : string
        Use a comma delimiter by default.

    names : boolean
        If `True`, the first row contains the field names.
    encoding :
        Set to None to use system default
    see np.genfromtxt for all *args and **kwargs.
    """
    a = np.genfromtxt(name, dtype=None, delimiter=",",
                      names=True, autostrip=True, encoding=None)  # ,skip_header=1)
    return a


# ----------------------------------------------------------------------
# (4) save_txt .... code section ---
def save_txt(a, name="arr.txt", sep=", ", dt_hdr=True):
    """Save a NumPy structured, recarray to text.

    Requires:
    --------
    a     : array
        input array
    fname : filename
        output filename and path otherwise save to script folder
    sep   : separator
        column separater, include a space if needed
    dt_hdr: boolean
        if True, add dtype names to the header of the file
    """
    a_names = ", ".join(i for i in a.dtype.names)
    hdr = ["", a_names][dt_hdr]  # use "" or names from input array
    s = np.array(a.tolist(), dtype=np.unicode_)
    widths = [max([len(i) for i in s[:, j]])
              for j in range(s.shape[1])]
    frmt = sep.join(["%{}s".format(i) for i in widths])
    # vals = ", ".join([i[1] for i in a.dtype.descr])
    np.savetxt(name, a, fmt=frmt, header=hdr, comments="")
    print("\nFile saved...")


# ----------------------------------------------------------------------
# (5) arr_json .... code section ---
def arr_json(file_out, arr=None):
    """Send an array out to json format. Use json_arr to read the file.
    No error checking
    """
    import json
    import codecs
    json.dump(arr.tolist(), codecs.open(file_out, 'w', encoding='utf-8'),
              sort_keys=True, indent=4)
    # ----


# ----------------------------------------------------------------------
# (6 - 9) Dictionary - array section
# dict_arrays, iterable_dict, struct_dict
def dict_arrays(d):
    """Dictionary to arrays

    Parameters:
    -----------
    d : dictionary
        The dictionary to convert to arrays

    Returns:
    --------
    A list, which can be converted to an array if needed.  It will probably
    be an `object` array if the array types are mixed.

    >>> d =  {'A': 1, 'B': [1, 2], 'C': (3.0, 4), 'D': (5.0, 6.0),
              'E': ['a', 'bb'], 'F': ['ccc', 7, 8.0], 'G': [[1, 2], [3, 4]]}
    >>> arr = np.asarray(dict_arrays(d))
    >>> arr
    array([array(1), array([1, 2]), array([3., 4.]), array([5., 6.]),
           array(['a', 'bb'], dtype='<U2'),
           array(['ccc', 7, 8.0], dtype=object),
           array([[1, 2], [3, 4]], dtype=object)], dtype=object)
    """
#    def dtstr(v):
    keys = d.keys()
    dts = []
    vals = []
    for k in keys:
        v = d[k]
        if isinstance(v, int):
            dts.append('<i4')
        elif isinstance(v, float):
            dts.append('<f8')
        if isinstance(v, (list, tuple, np.ndarray)):
            if all(isinstance(x, (int, float)) for x in v):
                dts.append(np.array(v).dtype.str)
            elif all(isinstance(x, (str)) for x in v):
                m = len(max(v, key=len))
                dts.append('<U' + str(m))
            else:
                dts.append('O')
        vals.append(v)
    arrs = [np.array(i[0], dtype=i[1]) for i in list(zip(vals, dts))]
    return arrs


def iterable_dict(a, use_numbers=True):
    """Iterable (list, tuple, np.ndarray) to dictionary

    Parameters:
    -----------
    a : iterable
        The iterable to convert to a dictionary.  Normally useful if it is a
        list of lists or np.ndarray
    use_numbers : boolean
        True, the dictionary keys are assigned 0...n.  False, letters are
        sliced from the lett_lst below

    Returns:
    --------
    A dictionary.

    >>> # see `arr` in `dict_arrays`
    >>> iterable_dict(arr, use_numbers=False)
    >>> {'A': array(1), 'B': array([1, 2]), 'C': array([3., 4.]),
         'D': array([5., 6.]), 'E': array(['a', 'bb'], dtype='<U2'),
         'F': array(['ccc', 7, 8.0], dtype=object),
         'G': array([[1, 2], [3, 4]], dtype=object)}

    Simpler cases

    >>> z = 'abcde'
    >>> iterable_dict(z)
    {0: 'a', 1: 'b', 2: 'c', 3: 'd', 4: 'e'}
    >>> iterable_dict(z, False)
    {'A': 'a', 'B': 'b', 'C': 'c', 'D': 'd', 'E': 'e'}
    >>> iterable_dict([1,'a', 2], False)
    {'A': 1, 'B': 'a', 'C': 2}

    """
    if use_numbers:
        d = {i: a[i] for i, e in enumerate(a)}
    else:
        lett_lst = list('ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz')
        d = {lett_lst[i]: a[i] for i, e in enumerate(a)}
    return d


def dict_struct(d):
    """Dictionary to simple structured/recarray
    """
    if not isinstance(d, dict):
        return d
    lens = [len(i) for i in d.values()]
    if max(lens) != min(lens):
        return d
    #
    names = [k for k in d.keys()]
    data = list(d.values())
    dt_str = [i.dtype.str for i in data]
    dt = list(zip(names, dt_str))
    N = lens[0]
    n = len(dt_str)
    arr = np.zeros((N,), dtype=dt)
    for i in range(n):
        arr[names[i]] = data[i]
    return arr


def struct_dict(a):
    """Structured/recarray to dictionary
    """
    if not isinstance(a, np.ndarray):
        return a
    if a.dtype.names is None:
        return a
    names = a.dtype.names
    return {i: a[i].tolist() for i in names}


# ----------------------------------------------------------------------
# (10) excel_np
def excel_np(path, sheet_num=0, int_null=-999):
    """Read excel files to numpy structured/record arrays.  Your spreadsheet
    must adhere to simple rules::
      - first row must contain the field names for the output array
      - no blank rows or columns, basically, no fluff or formatting
      - if you have nodata values, put them in, since blank cells will be
        'corrected' as best as possible.
      - text and numbers in a column, results in a text column

    Parameters:
    -----------
    path : text
        Full path to the xls, xlsx file
    sheet_num : integer
        Sheets are numbered from 0.

    int_null : integer
        Integer value to use for nulls. Strings have `None`, floats `np.nan`
        but integers have no equivalent so you have to provide one.
        you could use np.iinfo(np.intXX).min where XX is 8, 16, 32 to reflect
        the appropriate integer minimums

    Returns:
    --------
    A numpy structured array is returned.  Excel only uses float or string
    data, so attempts are made to coerse integer columns by comparing the
    float vs int versions of the arrays.  A tad of a kludge, but it works.

    The first row's data type is compared to its matching column data type.
    If they match, then it is used as the dtype.  If there is a mismatch an
    attempt is made to recover numeric data by assigning blanks etc in numeric
    columns a value of np.nan.

    String/text columns are check for empty cells, '', "" and that ever so
    ugly invisible space.

    Notes:
    ------
    >>> aString = open('c:/temp/test.xlsx','rb').read()
    >>> book_ = open_workbook(file_contents=aString)
    >>> dir(book_):
        get_sheet, nsheets, sheet_by_index, sheet_by_name etc....

    Now you can read a sheet

    >>> sheet = book_.sheet_by_index(0)  # first sheet
    >>> sheet.col_types(0)

    References:
    ----------
    `<https://media.readthedocs.org/pdf/xlrd/latest/xlrd.pdf>`_.
    """
    def _values(sheet, rows, cols):
        """return cell types for the above.  Skip the first row
        Not used .... just kept for future reference
        """
        ar = []
        for i in range(1, rows):
            c = []
            for j in range(cols):
                c.append(sheet.cell_values(i, j))  # sheet.cell_types also
            ar.append(c)
        return ar

    def isfloat(a):
        """float check"""
        try:
            i = float(a)
            return i
        except ValueError:
            return np.nan

    def punc_space(name):
        """delete punctuation and spaces and replace with '_'"""
        punc = list('!"#$%&\'()*+,-./:;<=>?@[\\]^`{|}~ ')
        return "".join([[i, '_'][i in punc] for i in name])

    import xlrd
    w = xlrd.open_workbook(path)        # xlrd.book.Book class
    sheet = w.sheet_by_index(sheet_num) # sheet by number
    # sheet = w.sheet_by_name('test')   # case sensitive, not implemented
    names = sheet.row_values(0)         # clean these up later
    cols = sheet.ncols
    rows = sheet.nrows
    col_data = [sheet.col_values(i, 1, rows) for i in range(cols)]
    row_guess = sheet.row_values(1)
    row_dts = [np.asarray(i).dtype.kind for i in row_guess]
    col_dts = [np.asarray(col_data[i]).dtype.kind
               for i in range(cols)]
    clean = []
    for i in range(len(row_dts)):
        c = col_data[i]
        if row_dts[i] == col_dts[i]:    # same dtype... send to array
            ar = np.asarray(c)
        if row_dts[i] == 'f':           # float? if so, substitute np.nan
            ar = np.array([isfloat(i) for i in c])
            is_nan = np.isnan(ar)       # find the nan values, then check
            not_nan = ar[~is_nan]       # are the floats == ints?
            if np.all(np.equal(not_nan, not_nan.astype('int'))):  # integer?
                ar[is_nan] = int_null   # assign the integer null
                ar = ar.astype('int')
        elif row_dts[i] in ('U', 'S'):  # unicode/string... send to array
            ar = np.char.strip(ar)
            ar = np.where(np.char.str_len(ar) == 0, 'None', ar)
        else:
            ar = np.asarray(c)
        clean.append(ar)
    # ---- assemble the columns for the array ----
    dt_str = [i.dtype.str for i in clean]
    names = [i.strip() for i in names]      # clean up leading/trailing spaces
    names = [punc_space(i) for i in names]  # replace punctuation and spaces
    dts_name = list(zip(names, dt_str))
    arr = np.empty((rows-1,), dtype=dts_name)
    cnt = 0
    for i in names:
        arr[i] = clean[cnt]
        cnt += 1
    return arr


def openxl_np(path):
    """read excel using openpyxl
    A = np.array([[i.value for i in j] for j in ws['C1':'E38']])

    `<https://stackoverflow.com/questions/35823835/reading-excel-file-is-
    magnitudes-slower-using-openpyxl-compared-to-xlrd>`_.

    cols = sheet.max_column  # sheet.min_column  1 to max
    rows = sheet.max_row  # sheet.min_row
    """
    import openpyxl as op
    wb = op.load_workbook(path, data_only=True, guess_types=True,
                          keep_links=False)
    sheets = wb.sheetnames
    sheet = wb[sheets[0]]
    data = list(sheet.values)
    #header = data[0]
    return data

def _demo_npy():
    """
    """
    _npy = "/Data/sample_10K.npy"  # change to one in the Data folder
    _npy = "{}".format(script.replace("a_io.py", _npy))
    return _npy

def _demo_xlsx():
    """
    """
#    _xlsx = "/Data/test2.xlsx"  # page 0 or page 1
    _xlsx = "/Data/Test_10K.xlsx"
    path = script.rpartition("/")[0] + _xlsx
    arr = excel_np(path, 0)
    return arr
# ----------------------------------------------------------------------
# __main__ .... code section
if __name__ == "__main__":
    """Optionally...
    : - print the script source name.
    : - run the _demo
    """
#    print("Script... {}".format(script))
#    fname = _demo()
